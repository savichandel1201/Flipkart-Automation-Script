import openpyxl

class ExcelUtils:
    @staticmethod
    def get_data_from_excel(file_path, sheet_name):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        data = []
        
        # Assume first row is header, data starts from row 2
        for row in range(2, sheet.max_row + 1):
            product = sheet.cell(row=row, column=1).value
            if product:
                data.append(product)
        return data

    @staticmethod
    def write_data_to_excel(file_path, sheet_name, row_num, col_num, data):
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        sheet.cell(row=row_num, column=col_num).value = data
        workbook.save(file_path)
